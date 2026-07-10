"""Validate and convert the NATS SRD Excel workbook to CSV files.

Workflow
--------
1. Open the workbook and locate the "What's New" sheet.
2. Scan its cells for the validation header — a cell whose text matches
   ``"What's New - {D}{ord} {Month} {YYYY} AIRAC"`` — and verify the embedded
   date matches the target AIRAC cycle's effective date.
3. If validation passes, export the "Routes" and "Notes" worksheets to CSV.

The validation step catches wrong-cycle downloads early and with a visible,
operator-readable error message before any downstream tool consumes bad data.

Sheet name and header conventions
----------------------------------
# [RULE:SRD-EXCEL-STRUCTURE]
# "What's New" — validation sheet; header cell contains the effective date.
# "Routes"     — route data exported to CSV for the SRD Parser.
# "Notes"      — supplementary notes exported to CSV for manual reference.
# Header format: "What's New - {D}{ord} {Month} {YYYY} AIRAC"
#   e.g. "What's New - 19th February 2026 AIRAC"
# The date must match cycle.effective_date exactly.
"""

from __future__ import annotations

import csv
import logging
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

from src.airac import AiracCycle

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants  [RULE:SRD-EXCEL-STRUCTURE]
# ---------------------------------------------------------------------------

_WHATS_NEW_SHEET = "What's New"
_ROUTES_SHEET = "Routes"
_NOTES_SHEET = "Notes"

# Regex matches: "What's New - 19th February 2026 AIRAC"
# Groups: (day, month_name, year)
_HEADER_RE = re.compile(
    r"What's\s+New\s*-\s*(\d{1,2})(?:st|nd|rd|th)\s+(\w+)\s+(\d{4})\s+AIRAC",
    re.IGNORECASE,
)

# How many rows / columns to scan in the "What's New" sheet for the header
_SCAN_ROWS = 30
_SCAN_COLS = 10


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class ExcelValidationError(Exception):
    """Raised when the SRD workbook fails structure or date validation."""


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _find_whats_new_date(ws) -> date | None:
    """Scan *ws* for the validation header and return the parsed date.

    Returns None if no matching cell is found within the scan window.
    # [RULE:SRD-EXCEL-STRUCTURE]
    """
    for row in ws.iter_rows(max_row=_SCAN_ROWS, max_col=_SCAN_COLS, values_only=True):
        for cell_value in row:
            if cell_value is None:
                continue
            normalised = str(cell_value).replace("\u2019", "'").replace("\u2018", "'")
            m = _HEADER_RE.search(normalised)
            if m:
                day, month_name, year = m.group(1), m.group(2), m.group(3)
                try:
                    return datetime.strptime(
                        f"{day} {month_name} {year}", "%d %B %Y"
                    ).date()
                except ValueError:
                    continue  # unparseable month name — keep scanning
    return None


def _last_populated_column(ws) -> int:
    """Return the count of columns up to and including the last one that
    holds a non-empty value anywhere in *ws*.

    Some workbooks report an inflated ``ws.dimensions``/``max_column`` far
    beyond the real data extent (e.g. stray formatting applied to a large
    range). Trusting that bound when exporting to CSV pads every row with
    thousands of empty trailing fields, bloating file size by orders of
    magnitude with no content gained. Scanning for the true rightmost
    non-empty cell keeps the export trimmed to the actual data.
    """
    max_col = 0
    for row in ws.iter_rows(values_only=True):
        for idx in range(len(row) - 1, max_col - 1, -1):
            value = row[idx]
            if value is not None and str(value).strip() != "":
                max_col = idx + 1
                break
    return max_col


def _sheet_to_csv(ws, dest_path: Path) -> Path:
    """Write all rows from *ws* to a CSV file at *dest_path*.

    None values are written as empty strings so the CSV remains well-formed.
    Rows are trimmed to the last column containing any value anywhere in the
    sheet, discarding trailing empty columns (see `_last_populated_column`).
    """
    max_col = _last_populated_column(ws)
    with dest_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        for row in ws.iter_rows(values_only=True):
            trimmed = row[:max_col] if max_col else row
            writer.writerow(["" if v is None else v for v in trimmed])
    return dest_path


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def validate_whats_new(wb, cycle: AiracCycle) -> None:
    """Verify the "What's New" sheet date matches *cycle*.

    Raises ExcelValidationError with an operator-readable message on any failure.
    # [RULE:SRD-EXCEL-STRUCTURE]
    """
    if _WHATS_NEW_SHEET not in wb.sheetnames:
        raise ExcelValidationError(
            f"SRD workbook is missing the '{_WHATS_NEW_SHEET}' worksheet.\n"
            "The workbook structure may have changed. [RULE:SRD-EXCEL-STRUCTURE]"
        )

    ws = wb[_WHATS_NEW_SHEET]
    found_date = _find_whats_new_date(ws)

    if found_date is None:
        raise ExcelValidationError(
            f"Could not find the validation header in the '{_WHATS_NEW_SHEET}' sheet.\n"
            "Expected a cell containing: "
            "\"What's New - {D}th {Month} {YYYY} AIRAC\"\n"
            "The workbook format may have changed. [RULE:SRD-EXCEL-STRUCTURE]"
        )

    if found_date != cycle.effective_date:
        raise ExcelValidationError(
            f"SRD workbook date mismatch for cycle {cycle.ident}!\n"
            f"  '{_WHATS_NEW_SHEET}' sheet says: {found_date.isoformat()}\n"
            f"  Expected (cycle effective date): {cycle.effective_date.isoformat()}\n"
            "This file may be for the wrong AIRAC cycle. "
            "Check the download and try again. [RULE:SRD-EXCEL-STRUCTURE]"
        )


def convert_srd_excel(
    excel_path: Path,
    dest_dir: Path,
    cycle: AiracCycle,
) -> dict[str, Path]:
    """Validate the SRD Excel workbook and export Routes and Notes to CSV.

    Args:
        excel_path: Path to the downloaded SRD .xlsx file.
        dest_dir:   Directory where CSV files will be written.
                    Created if it does not already exist.
        cycle:      The target AIRAC cycle (used for date validation).

    Returns:
        A dict mapping sheet name -> CSV Path:
        ``{"Routes": Path("…/Routes.csv"), "Notes": Path("…/Notes.csv")}``

    Raises:
        ExcelValidationError: if the workbook fails structure or date checks.
    """
    logger.info("Opening SRD workbook: %s", excel_path.name)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    logger.info("Sheets found: %s", wb.sheetnames)

    # Step 1: validate the What's New sheet before touching any data
    validate_whats_new(wb, cycle)
    logger.info("What's New date validated for cycle %s", cycle.ident)

    dest_dir.mkdir(parents=True, exist_ok=True)
    result: dict[str, Path] = {}

    # Step 2: export Routes and Notes
    # [RULE:SRD-EXCEL-STRUCTURE]
    for sheet_name in (_ROUTES_SHEET, _NOTES_SHEET):
        if sheet_name not in wb.sheetnames:
            raise ExcelValidationError(
                f"SRD workbook is missing the '{sheet_name}' worksheet.\n"
                "The workbook structure may have changed. [RULE:SRD-EXCEL-STRUCTURE]"
            )
        csv_path = dest_dir / f"{sheet_name}.csv"
        _sheet_to_csv(wb[sheet_name], csv_path)
        result[sheet_name] = csv_path
        logger.info("Exported %s → %s (%d bytes)", sheet_name, csv_path.name, csv_path.stat().st_size)

    return result
