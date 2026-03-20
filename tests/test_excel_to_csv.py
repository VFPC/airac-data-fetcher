"""Unit tests for src/processing/excel_to_csv.py.

Tests build real openpyxl workbooks in tmp_path — no mocking of openpyxl.
All filesystem writes go through pytest's tmp_path fixture.
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from src.airac import AiracCycle
from src.processing.excel_to_csv import (
    ExcelValidationError,
    _NOTES_SHEET,
    _ROUTES_SHEET,
    _WHATS_NEW_SHEET,
    _find_whats_new_date,
    convert_srd_excel,
    validate_whats_new,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

CYCLE_2602 = AiracCycle(
    year=2026,
    number=2,
    effective_date=date(2026, 2, 19),
    expiry_date=date(2026, 3, 18),
)

CYCLE_2601 = AiracCycle(
    year=2026,
    number=1,
    effective_date=date(2026, 1, 22),
    expiry_date=date(2026, 2, 18),
)

# Confirmed production header format [RULE:SRD-EXCEL-STRUCTURE]
VALID_HEADER = "What's New - 19th February 2026 AIRAC"

ROUTES_DATA = [
    ("Route", "From", "To", "Level", "Direction"),
    ("UL9", "REMSI", "DVR", "FL245", "Both"),
    ("UM605", "GAPLI", "BRAIN", "FL265", "Both"),
]

NOTES_DATA = [
    ("Note", "Details"),
    ("Amendment", "Some route change"),
]


def _make_workbook(
    tmp_path: Path,
    whats_new_header: str = VALID_HEADER,
    whats_new_cell: str = "A1",
    routes_data: list | None = None,
    notes_data: list | None = None,
    omit_sheets: list[str] | None = None,
) -> Path:
    """Build a minimal SRD-like workbook and return its path."""
    wb = openpyxl.Workbook()
    # openpyxl always creates a default "Sheet" — remove it
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    omit = set(omit_sheets or [])

    if _WHATS_NEW_SHEET not in omit:
        ws_wn = wb.create_sheet(_WHATS_NEW_SHEET)
        if whats_new_header:
            ws_wn[whats_new_cell] = whats_new_header

    if _ROUTES_SHEET not in omit:
        ws_routes = wb.create_sheet(_ROUTES_SHEET)
        for row in (routes_data or ROUTES_DATA):
            ws_routes.append(row)

    if _NOTES_SHEET not in omit:
        ws_notes = wb.create_sheet(_NOTES_SHEET)
        for row in (notes_data or NOTES_DATA):
            ws_notes.append(row)

    path = tmp_path / "SRD.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# _find_whats_new_date
# ---------------------------------------------------------------------------

class TestFindWhatsNewDate:
    def _ws_with_cell(self, value: str, cell: str = "A1"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws[cell] = value
        return ws

    def test_finds_standard_header(self):
        # [RULE:SRD-EXCEL-STRUCTURE]
        ws = self._ws_with_cell(VALID_HEADER)
        assert _find_whats_new_date(ws) == date(2026, 2, 19)

    def test_finds_first_of_month(self):
        ws = self._ws_with_cell("What's New - 1st January 2026 AIRAC")
        assert _find_whats_new_date(ws) == date(2026, 1, 1)

    def test_finds_second_ordinal(self):
        ws = self._ws_with_cell("What's New - 22nd January 2026 AIRAC")
        assert _find_whats_new_date(ws) == date(2026, 1, 22)

    def test_finds_third_ordinal(self):
        ws = self._ws_with_cell("What's New - 3rd March 2025 AIRAC")
        assert _find_whats_new_date(ws) == date(2025, 3, 3)

    def test_case_insensitive(self):
        ws = self._ws_with_cell("WHAT'S NEW - 19TH FEBRUARY 2026 AIRAC")
        assert _find_whats_new_date(ws) == date(2026, 2, 19)

    def test_header_not_in_a1_still_found(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Standard Route Document"
        ws["A3"] = "Effective date information"
        ws["B5"] = VALID_HEADER
        assert _find_whats_new_date(ws) == date(2026, 2, 19)

    def test_no_matching_cell_returns_none(self):
        ws = self._ws_with_cell("This is just a title")
        assert _find_whats_new_date(ws) is None

    def test_empty_sheet_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert _find_whats_new_date(ws) is None

    def test_header_embedded_in_longer_string(self):
        ws = self._ws_with_cell(f"  {VALID_HEADER}  extra text")
        assert _find_whats_new_date(ws) == date(2026, 2, 19)

    def test_unicode_right_single_quotation_mark(self):
        # NATS uses U+2019 (RIGHT SINGLE QUOTATION MARK) in some cycles
        ws = self._ws_with_cell("What\u2019s New - 19th February 2026 AIRAC")
        assert _find_whats_new_date(ws) == date(2026, 2, 19)

    def test_header_with_version_suffix(self):
        # NATS appends "(v2)" or similar in revised workbooks
        ws = self._ws_with_cell("What\u2019s New - 19th March 2026 AIRAC (v2)")
        assert _find_whats_new_date(ws) == date(2026, 3, 19)


# ---------------------------------------------------------------------------
# validate_whats_new
# ---------------------------------------------------------------------------

class TestValidateWhatsNew:
    def test_valid_workbook_passes(self, tmp_path):
        path = _make_workbook(tmp_path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        validate_whats_new(wb, CYCLE_2602)  # must not raise

    def test_date_mismatch_raises(self, tmp_path):
        path = _make_workbook(tmp_path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        with pytest.raises(ExcelValidationError, match="date mismatch"):
            validate_whats_new(wb, CYCLE_2601)

    def test_mismatch_message_shows_both_dates(self, tmp_path):
        path = _make_workbook(tmp_path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        with pytest.raises(ExcelValidationError) as exc_info:
            validate_whats_new(wb, CYCLE_2601)
        msg = str(exc_info.value)
        assert "2026-02-19" in msg  # what the file says
        assert "2026-01-22" in msg  # what was expected

    def test_mismatch_message_shows_cycle_ident(self, tmp_path):
        path = _make_workbook(tmp_path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        with pytest.raises(ExcelValidationError, match="2601"):
            validate_whats_new(wb, CYCLE_2601)

    def test_missing_whats_new_sheet_raises(self, tmp_path):
        path = _make_workbook(tmp_path, omit_sheets=[_WHATS_NEW_SHEET])
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        with pytest.raises(ExcelValidationError, match=_WHATS_NEW_SHEET):
            validate_whats_new(wb, CYCLE_2602)

    def test_header_not_found_raises(self, tmp_path):
        path = _make_workbook(tmp_path, whats_new_header="")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        with pytest.raises(ExcelValidationError, match="validation header"):
            validate_whats_new(wb, CYCLE_2602)

    def test_rule_tag_in_error_message(self, tmp_path):
        path = _make_workbook(tmp_path, omit_sheets=[_WHATS_NEW_SHEET])
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        with pytest.raises(ExcelValidationError, match="RULE:SRD-EXCEL-STRUCTURE"):
            validate_whats_new(wb, CYCLE_2602)


# ---------------------------------------------------------------------------
# convert_srd_excel
# ---------------------------------------------------------------------------

class TestConvertSrdExcel:
    def test_returns_routes_and_notes_paths(self, tmp_path):
        path = _make_workbook(tmp_path)
        result = convert_srd_excel(path, tmp_path / "out", CYCLE_2602)
        assert set(result.keys()) == {_ROUTES_SHEET, _NOTES_SHEET}

    def test_csv_files_exist(self, tmp_path):
        out = tmp_path / "out"
        path = _make_workbook(tmp_path)
        result = convert_srd_excel(path, out, CYCLE_2602)
        for p in result.values():
            assert p.exists()

    def test_routes_csv_filename(self, tmp_path):
        out = tmp_path / "out"
        path = _make_workbook(tmp_path)
        result = convert_srd_excel(path, out, CYCLE_2602)
        assert result[_ROUTES_SHEET].name == "Routes.csv"

    def test_notes_csv_filename(self, tmp_path):
        out = tmp_path / "out"
        path = _make_workbook(tmp_path)
        result = convert_srd_excel(path, out, CYCLE_2602)
        assert result[_NOTES_SHEET].name == "Notes.csv"

    def test_routes_csv_contains_data(self, tmp_path):
        out = tmp_path / "out"
        path = _make_workbook(tmp_path)
        result = convert_srd_excel(path, out, CYCLE_2602)
        rows = list(csv.reader(result[_ROUTES_SHEET].open(encoding="utf-8")))
        # Header row
        assert rows[0] == list(ROUTES_DATA[0])
        # Data row
        assert rows[1] == list(ROUTES_DATA[1])

    def test_notes_csv_contains_data(self, tmp_path):
        out = tmp_path / "out"
        path = _make_workbook(tmp_path)
        result = convert_srd_excel(path, out, CYCLE_2602)
        rows = list(csv.reader(result[_NOTES_SHEET].open(encoding="utf-8")))
        assert rows[0] == list(NOTES_DATA[0])

    def test_dest_dir_created_if_missing(self, tmp_path):
        out = tmp_path / "new" / "subdir"
        assert not out.exists()
        path = _make_workbook(tmp_path)
        convert_srd_excel(path, out, CYCLE_2602)
        assert out.is_dir()

    def test_none_cells_become_empty_string(self, tmp_path):
        out = tmp_path / "out"
        sparse_data = [("A", None, "C"), (None, "B", None)]
        path = _make_workbook(tmp_path, routes_data=sparse_data)
        result = convert_srd_excel(path, out, CYCLE_2602)
        rows = list(csv.reader(result[_ROUTES_SHEET].open(encoding="utf-8")))
        assert rows[0] == ["A", "", "C"]
        assert rows[1] == ["", "B", ""]

    def test_date_mismatch_raises_before_writing_csv(self, tmp_path):
        out = tmp_path / "out"
        path = _make_workbook(tmp_path)
        with pytest.raises(ExcelValidationError, match="date mismatch"):
            convert_srd_excel(path, out, CYCLE_2601)
        # CSV files must NOT have been written
        assert not (out / "Routes.csv").exists()
        assert not (out / "Notes.csv").exists()

    def test_missing_routes_sheet_raises(self, tmp_path):
        out = tmp_path / "out"
        path = _make_workbook(tmp_path, omit_sheets=[_ROUTES_SHEET])
        with pytest.raises(ExcelValidationError, match=_ROUTES_SHEET):
            convert_srd_excel(path, out, CYCLE_2602)

    def test_missing_notes_sheet_raises(self, tmp_path):
        out = tmp_path / "out"
        path = _make_workbook(tmp_path, omit_sheets=[_NOTES_SHEET])
        with pytest.raises(ExcelValidationError, match=_NOTES_SHEET):
            convert_srd_excel(path, out, CYCLE_2602)

    def test_csv_files_in_dest_dir(self, tmp_path):
        out = tmp_path / "out"
        path = _make_workbook(tmp_path)
        result = convert_srd_excel(path, out, CYCLE_2602)
        for p in result.values():
            assert p.parent == out
